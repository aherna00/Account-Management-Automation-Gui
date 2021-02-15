from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl import Workbook
import time
from tkinter import *
from tkinter import filedialog

fidelity_accounts = []
funded_accounts = []
saveloc = ""


def selenium():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "funded accounts"
    driver = webdriver.Chrome()
    driver.implicitly_wait(60)
    driver.get("https://www.wealthscape.com/")
    wspass = e.get()
    user_name = driver.find_element_by_id("userInput")
    user_name.send_keys("*********")

    driver.find_element_by_class_name("group-h__item-spaced").click()
    driver.find_element_by_id("password").send_keys(wspass)

    driver.find_element_by_id("fs-login-button").click()

    for i in fidelity_accounts:
        driver.find_element_by_xpath \
                (
                "//ui-view/fip-basis-layout-portal-container/div/div/div/fip-side-bar-menu/bss-side-menu/div[1]/button").click()
        look_up = driver.find_element_by_id("ual-search-primary-input-unmasked")
        look_up.clear()
        look_up.send_keys(i)
        time.sleep(1)
        search_click = driver.find_element_by_xpath \
                (
                "/html/body/div/ui-view/fip-basis-layout-portal-container/div/div/div/fip-side-bar-item-panel/div[1]/div[2]/fip-unified-access-layer/unified-access-layer/div/unified-access-layer/div/div[1]/unified-access-layer-search-bar-sidebar/div/div/div/div/div/div[1]/div/span[5]/button")
        search_click.click()
        time.sleep(3)
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH,
                                            "/html/body/div/ui-view/fip-basis-layout-portal-container/div/div/div/main/ui-view[2]/ui-view/fip-basis-layout-account-page/div/div[1]/account-cartridge/div/div[2]/div/extend-cartridge-details/section[2]/div[1]/div[1]/ul/li[1]/div/span"))
        )
        balance = driver.find_element_by_xpath(
            "/html/body/div/ui-view/fip-basis-layout-portal-container/div/div/div/main/ui-view[2]/ui-view/fip-basis-layout-account-page/div/div[1]/account-cartridge/div/div[2]/div/extend-cartridge-details/section[2]/div[1]/div[1]/ul/li[1]/div/span").get_attribute(
            "innerHTML")

        if balance != "$0.00":
            print(i + " is funded " + balance)
            funded_accounts.append(i + " is funded " + balance)
        else:
            pass

    column = 1
    for r, value in enumerate(funded_accounts):
        ws1.cell(column=column, row=r+1, value=value)

    wb.save(saveloc)
    root.destroy()
    driver.close()


def excelpath():
    filename = filedialog.askopenfilename(title="Select Excel",
                                          filetypes=[("excel files", "*.xlsx")])
    path = filename
    Label(root, text="File Location:").pack()
    Label(root, text=filename).pack()
    read_acc = load_workbook(str(path))

    sheet = read_acc.active
    m_row = sheet.max_row

    print(m_row)
    # used for debugging

    for i in range(1, m_row + 1):
        cell_acc = sheet.cell(row=i, column=1)
        fidelity_accounts.append(cell_acc.value)


def savepath():
    savep = filedialog.askdirectory()
    Label(root, text="Save Location:").pack()
    Label(root, text=savep).pack()
    global saveloc
    saveloc = savep + "/funded.xlsx"


root = Tk()
root.geometry("400x400")

myLabel = Label(root, text="Funded Account Tool", font="Helvetica")
myLabel.pack()

myButton = Button(root, text="Select Excel file", command=lambda: excelpath())
myButton.pack(pady=10)

saveButton = Button(root, text="Select Save Location", command=lambda: savepath())
saveButton.pack(pady=10)

passwordLabel = Label(root, text="Enter Wealthscape Password:")
passwordLabel.pack(pady=10)

e = Entry(root, show="*")
e.pack()

runButton = Button(root, text="Start", command=selenium)
runButton.pack(pady=30)

createdByLabel = Label(root, text="Developed by Alex Hernandez")
createdByLabel.pack(pady=20)

root.mainloop()
